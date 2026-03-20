"""build_comparison.py — Add a Model Comparison sheet to ufc_bet_companion_picks.xlsx.

Matches our past_predictions (DB) against UFC Bet Companion's historic AI picks,
side by side per fight with P&L computed using UFC BC's odds at prediction.

Usage:
    cd backend
    python eda/build_comparison.py
"""
from __future__ import annotations

import os, sys
sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, os.path.abspath(os.path.dirname(__file__) + "/.."))

from sqlalchemy import text
from db.database import engine

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from rapidfuzz import fuzz
from pathlib import Path

XLSX = Path(__file__).parent / "ufc_bet_companion_picks.xlsx"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _name_key(name: str) -> str:
    """Normalise fighter name for matching."""
    return name.lower().strip().replace(".", "").replace("-", " ").replace("'", "")


def _match_score(our_a: str, our_b: str, bc_fight: str) -> float:
    """Score how well our (fighter_a, fighter_b) matches a 'X vs Y' string."""
    parts = bc_fight.lower().split(" vs ")
    if len(parts) != 2:
        return 0
    bc_a, bc_b = parts[0].strip(), parts[1].strip()
    fwd = (fuzz.token_sort_ratio(_name_key(our_a), _name_key(bc_a)) +
           fuzz.token_sort_ratio(_name_key(our_b), _name_key(bc_b))) / 2
    rev = (fuzz.token_sort_ratio(_name_key(our_a), _name_key(bc_b)) +
           fuzz.token_sort_ratio(_name_key(our_b), _name_key(bc_a))) / 2
    return max(fwd, rev)


def _pnl_for_winner(is_correct: bool | None, odds: int | None, profit: float | None) -> float | None:
    """
    Compute flat-$100 P&L for the given predicted winner.
    odds: American odds for that fighter.
    """
    if is_correct is None or odds is None:
        return None
    if is_correct:
        return round(100 / abs(odds) * 100, 2) if odds < 0 else round(odds / 100 * 100, 2)
    return -100.0


def _mirror_odds(odds: int) -> int:
    """
    Estimate fair American odds for the opponent given their opponent's odds.
    Uses the zero-vig implied probability.
    e.g. -300 → ~+300, -150 → ~+150, +200 → ~-200
    """
    if odds < 0:
        prob_picked = abs(odds) / (abs(odds) + 100)
    else:
        prob_picked = 100 / (odds + 100)
    prob_other = 1 - prob_picked
    if prob_other >= 0.5:
        return round(-prob_other / (1 - prob_other) * 100)
    else:
        return round((1 - prob_other) / prob_other * 100)


# ---------------------------------------------------------------------------
# Load our past_predictions from DB
# ---------------------------------------------------------------------------

def load_our_predictions() -> list[dict]:
    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT DISTINCT ON (fight_id)
                fighter_a_id, fighter_b_id,
                fighter_a_name, fighter_b_name,
                event_name, event_date,
                win_prob_a, win_prob_b,
                predicted_winner_id, actual_winner_id,
                is_correct, confidence,
                pred_method_ko_tko, pred_method_sub, pred_method_dec,
                prediction_source
            FROM past_predictions
            WHERE event_date >= '2025-03-01'
            ORDER BY fight_id, prediction_source DESC
        """)).mappings().all()
    return [dict(r) for r in rows]


# ---------------------------------------------------------------------------
# Build comparison
# ---------------------------------------------------------------------------

def build_comparison(our_preds: list[dict], bc_rows: list[tuple]) -> list[dict]:
    """
    For each UFC BC fight, find the best-matching our prediction.
    Returns a list of comparison dicts.
    """
    MATCH_THRESHOLD = 75

    # Index our predictions by (fighter_a_name, fighter_b_name)
    unmatched_ours = list(our_preds)
    comparison: list[dict] = []
    matched_our_idxs: set[int] = set()

    for bc in bc_rows:
        # bc columns (from Fights sheet):
        # 0=Event, 1=EventAcc, 2=EventProfit, 3=FighterA, 4=FighterB,
        # 5=PredictedWinner, 6=Correct, 7=Confidence, 8=Odds, 9=Profit, 10=FightURL
        bc_event   = bc[0]
        bc_fight   = f"{bc[3]} vs {bc[4]}"
        bc_winner  = bc[5]
        bc_correct = bc[6]   # '✓' / '✗' / '?'
        bc_conf    = bc[7]
        bc_odds    = bc[8]
        bc_profit  = bc[9]

        # Find best match
        best_idx, best_score = -1, 0
        for i, pred in enumerate(our_preds):
            if i in matched_our_idxs:
                continue
            score = _match_score(
                pred["fighter_a_name"] or "",
                pred["fighter_b_name"] or "",
                bc_fight,
            )
            if score > best_score:
                best_score, best_idx = score, i

        our = our_preds[best_idx] if best_idx >= 0 and best_score >= MATCH_THRESHOLD else None
        if our:
            matched_our_idxs.add(best_idx)

        # Determine our predicted winner name
        our_pred_winner_name = None
        our_win_pct = None
        our_conviction = None
        our_correct_str = None
        our_pnl = None

        if our and our["win_prob_a"] is not None:
            if our["win_prob_a"] >= our["win_prob_b"]:
                our_pred_winner_name = our["fighter_a_name"]
                our_win_pct = our["win_prob_a"]
            else:
                our_pred_winner_name = our["fighter_b_name"]
                our_win_pct = our["win_prob_b"]
            our_conviction = round((our["win_pct"] if False else (max(our["win_prob_a"], our["win_prob_b"]) - 0.5) * 2) * 100, 1)

            our_is_correct = our["is_correct"]
            our_correct_str = "✓" if our_is_correct is True else ("✗" if our_is_correct is False else "·")

            # P&L using UFC BC's odds
            if bc_odds is not None and our["actual_winner_id"] is not None:
                # Did our predicted winner actually win?
                our_won = our_is_correct  # True/False/None
                # Were we picking the same fighter as UFC BC?
                same_pick = (
                    _name_key(our_pred_winner_name or "") in _name_key(bc_winner or "") or
                    _name_key(bc_winner or "") in _name_key(our_pred_winner_name or "") or
                    fuzz.token_sort_ratio(_name_key(our_pred_winner_name or ""), _name_key(bc_winner or "")) >= 80
                )
                if same_pick:
                    our_pnl = bc_profit  # same bet, same outcome
                else:
                    # We're betting the other fighter — estimate their odds
                    mirror = _mirror_odds(bc_odds)
                    our_pnl = _pnl_for_winner(our_won, mirror, None)
                    if our_pnl is not None:
                        our_pnl = round(our_pnl, 2)

        elif our:
            our_correct_str = "·"  # no prediction

        # Agreement
        if our_pred_winner_name and bc_winner:
            agree = (fuzz.token_sort_ratio(_name_key(our_pred_winner_name), _name_key(bc_winner)) >= 80)
        else:
            agree = None

        comparison.append({
            "event_date":      our["event_date"] if our else None,
            "event_name":      bc_event,
            "fight":           bc_fight,
            "our_winner":      our_pred_winner_name,
            "our_win_pct":     round(our_win_pct * 100, 1) if our_win_pct else None,
            "our_conviction":  our_conviction,
            "bc_winner":       bc_winner,
            "bc_confidence":   bc_conf,
            "agree":           agree,
            "actual_winner":   (our["fighter_a_name"] if our and our.get("actual_winner_id") == our.get("fighter_a_id")
                                 else our["fighter_b_name"] if our and our.get("actual_winner_id") == our.get("fighter_b_id")
                                 else None),
            "our_correct":     our_correct_str,
            "bc_correct":      bc_correct,
            "bc_odds":         bc_odds,
            "bc_pnl":          bc_profit,
            "our_pnl":         our_pnl,
            "match_score":     round(best_score, 1) if our else 0,
            "no_pred":         (our is not None and our["win_prob_a"] is None),
            "unmatched":       (our is None),
        })

    return comparison


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

HDR_FILL   = PatternFill("solid", fgColor="2D3748")
HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
GREEN_FONT = Font(color="166534", bold=True)   # dark green — readable on white
RED_FONT   = Font(color="991B1B", bold=True)   # dark red
GRAY_FONT  = Font(color="6B7280")
AMBER_FONT = Font(color="92400E", bold=True)
DARK_FONT  = Font(color="111827")
AGREE_FILL = PatternFill("solid", fgColor="DCFCE7")   # very light green
DISAGR_FILL= PatternFill("solid", fgColor="FEE2E2")   # very light red
NOPRED_FILL= PatternFill("solid", fgColor="F3F4F6")   # light gray


def _set_col_widths(ws, widths: dict[int, float]):
    for col, w in widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w


def _write_summary(ws, comparison: list[dict], start_row: int) -> int:
    """Write summary stats block. Returns the next free row."""
    matched = [c for c in comparison if not c["unmatched"]]
    both_pred = [c for c in matched if not c["no_pred"] and c["bc_correct"] in ("✓", "✗")]
    agrees = [c for c in both_pred if c["agree"] is True]
    disagrees = [c for c in both_pred if c["agree"] is False]

    our_correct   = sum(1 for c in both_pred if c["our_correct"] == "✓")
    bc_correct    = sum(1 for c in both_pred if c["bc_correct"] == "✓")
    agree_correct = sum(1 for c in agrees if c["our_correct"] == "✓")
    disagr_our_correct = sum(1 for c in disagrees if c["our_correct"] == "✓")
    disagr_bc_correct  = sum(1 for c in disagrees if c["bc_correct"] == "✓")

    our_pnl_sum = sum(c["our_pnl"] for c in both_pred if c["our_pnl"] is not None)
    bc_pnl_sum  = sum(c["bc_pnl"] for c in both_pred if c["bc_pnl"] is not None)

    stats = [
        ("Matched fights (both datasets)", len(matched)),
        ("Fights with predictions (both)", len(both_pred)),
        ("", ""),
        ("Our accuracy",    f"{our_correct}/{len(both_pred)} = {our_correct/len(both_pred)*100:.1f}%" if both_pred else "n/a"),
        ("UFC BC accuracy", f"{bc_correct}/{len(both_pred)} = {bc_correct/len(both_pred)*100:.1f}%" if both_pred else "n/a"),
        ("", ""),
        ("Agreements",  f"{len(agrees)} fights ({len(agrees)/len(both_pred)*100:.0f}%)  →  accuracy when we agree: {agree_correct}/{len(agrees)} = {agree_correct/len(agrees)*100:.1f}%" if agrees else "n/a"),
        ("Disagreements", f"{len(disagrees)} fights ({len(disagrees)/len(both_pred)*100:.0f}%)"),
        ("  Our accuracy on disagreements",  f"{disagr_our_correct}/{len(disagrees)} = {disagr_our_correct/len(disagrees)*100:.1f}%" if disagrees else "n/a"),
        ("  UFC BC accuracy on disagreements", f"{disagr_bc_correct}/{len(disagrees)} = {disagr_bc_correct/len(disagrees)*100:.1f}%" if disagrees else "n/a"),
        ("", ""),
        ("Our total P&L ($100/fight)",    f"${our_pnl_sum:+,.2f}  (ROI: {our_pnl_sum/(len(both_pred)*100)*100:.2f}%)"),
        ("UFC BC total P&L ($100/fight)", f"${bc_pnl_sum:+,.2f}  (ROI: {bc_pnl_sum/(len(both_pred)*100)*100:.2f}%)"),
        ("", ""),
        ("Note: 'Our P&L' uses UFC BC's odds. When picks differ, opponent odds are estimated (zero-vig mirror).", ""),
    ]

    ws.cell(start_row, 1, "=== HEAD-TO-HEAD SUMMARY ===").font = Font(bold=True, color="E63946", size=12)
    r = start_row + 1
    for label, val in stats:
        ws.cell(r, 1, label).font = Font(bold=bool(label and not label.startswith(" ")))
        ws.cell(r, 2, val)
        r += 1
    return r + 1


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("Loading our predictions from DB...")
    our_preds = load_our_predictions()
    print(f"  {len(our_preds)} fights loaded")

    wb = openpyxl.load_workbook(XLSX)
    bc_fights = list(wb["Fights"].iter_rows(min_row=2, values_only=True))
    print(f"  {len(bc_fights)} UFC BC fights loaded from Excel")

    print("Matching fights...")
    comparison = build_comparison(our_preds, bc_fights)

    matched   = sum(1 for c in comparison if not c["unmatched"])
    unmatched = sum(1 for c in comparison if c["unmatched"])
    print(f"  Matched: {matched}  |  Unmatched: {unmatched}")

    # Remove old sheet if exists
    if "Comparison" in wb.sheetnames:
        del wb["Comparison"]

    ws = wb.create_sheet("Comparison")

    # ---- Header row ----
    headers = [
        "Date", "Event", "Fight",
        "Our Pick", "Our Win %", "Our Conviction %",
        "BC Pick", "BC Confidence",
        "Agree?",
        "Actual Winner",
        "Our ✓/✗", "BC ✓/✗",
        "Odds (BC pick)", "BC P&L ($100)", "Our P&L ($100)",
        "Notes",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # ---- Data rows ----
    for c in comparison:
        agree_str = "✓ Agree" if c["agree"] is True else ("✗ Differ" if c["agree"] is False else "–")

        notes = ""
        if c["unmatched"]:  notes = "unmatched"
        elif c["no_pred"]:  notes = "no model prediction"

        row = [
            c["event_date"],
            c["event_name"],
            c["fight"],
            c["our_winner"]    or "–",
            c["our_win_pct"],
            c["our_conviction"],
            c["bc_winner"]     or "–",
            c["bc_confidence"],
            agree_str,
            c["actual_winner"] or "–",
            c["our_correct"]   or "–",
            c["bc_correct"]    or "–",
            c["bc_odds"],
            c["bc_pnl"],
            c["our_pnl"],
            notes,
        ]
        ws.append(row)
        rn = ws.max_row

        # Row background
        if c["unmatched"] or c["no_pred"]:
            fill = NOPRED_FILL
        elif c["agree"] is True:
            fill = AGREE_FILL
        elif c["agree"] is False:
            fill = DISAGR_FILL
        else:
            fill = None

        if fill:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(rn, col_idx).fill = fill

        # Correct indicators
        for col_idx, val in [(11, c["our_correct"]), (12, c["bc_correct"])]:
            cell = ws.cell(rn, col_idx)
            if val == "✓":  cell.font = GREEN_FONT
            elif val == "✗": cell.font = RED_FONT
            elif val == "·": cell.font = GRAY_FONT

        # Agree cell
        agree_cell = ws.cell(rn, 9)
        if c["agree"] is True:   agree_cell.font = Font(color="166534", bold=True)
        elif c["agree"] is False: agree_cell.font = Font(color="991B1B", bold=True)

        # P&L cells — dark text so readable on light row backgrounds
        for col_idx in [14, 15]:
            cell = ws.cell(rn, col_idx)
            val = cell.value
            if val is not None:
                cell.font = Font(color="166534", bold=True) if val > 0 else (Font(color="991B1B", bold=True) if val < 0 else GRAY_FONT)

    # ---- Summary block below data ----
    gap_row = ws.max_row + 3
    _write_summary(ws, comparison, gap_row)

    # ---- Column widths ----
    _set_col_widths(ws, {
        1: 12,  # Date
        2: 38,  # Event
        3: 38,  # Fight
        4: 24,  # Our pick
        5: 11,  # Our win %
        6: 16,  # Our conviction
        7: 24,  # BC pick
        8: 13,  # BC confidence
        9: 11,  # Agree?
        10: 24, # Actual winner
        11: 9,  # Our correct
        12: 9,  # BC correct
        13: 14, # Odds
        14: 14, # BC P&L
        15: 14, # Our P&L
        16: 22, # Notes
    })

    # Freeze top row
    ws.freeze_panes = "A2"

    wb.save(XLSX)
    print(f"\nSaved Comparison sheet -> {XLSX}")

    # Quick terminal summary
    both = [c for c in comparison if not c["unmatched"] and not c["no_pred"] and c["bc_correct"] in ("✓","✗")]
    agrees   = [c for c in both if c["agree"] is True]
    disagrees= [c for c in both if c["agree"] is False]
    our_acc  = sum(1 for c in both if c["our_correct"] == "✓") / len(both) * 100 if both else 0
    bc_acc   = sum(1 for c in both if c["bc_correct"] == "✓")  / len(both) * 100 if both else 0
    our_pnl  = sum(c["our_pnl"] for c in both if c["our_pnl"] is not None)
    bc_pnl   = sum(c["bc_pnl"]  for c in both if c["bc_pnl"] is not None)

    print(f"\n{'='*55}")
    print(f"  Matched fights with both predictions: {len(both)}")
    print(f"  Our accuracy:    {our_acc:.1f}%")
    print(f"  UFC BC accuracy: {bc_acc:.1f}%")
    print(f"  Agreements:   {len(agrees)}   Disagreements: {len(disagrees)}")
    if disagrees:
        d_our = sum(1 for c in disagrees if c["our_correct"] == "✓") / len(disagrees) * 100
        d_bc  = sum(1 for c in disagrees if c["bc_correct"] == "✓")  / len(disagrees) * 100
        print(f"  On disagreements → ours: {d_our:.1f}%  |  BC: {d_bc:.1f}%")
    print(f"  Our P&L:    ${our_pnl:+,.2f}  (ROI: {our_pnl/(len(both)*100)*100:+.2f}%)")
    print(f"  UFC BC P&L: ${bc_pnl:+,.2f}  (ROI: {bc_pnl/(len(both)*100)*100:+.2f}%)")
    print(f"{'='*55}")


if __name__ == "__main__":
    main()
