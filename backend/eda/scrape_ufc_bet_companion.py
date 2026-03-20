"""scrape_ufc_bet_companion.py — Scrape historic AI picks from ufcbetcompanion.com/ai-picks.

Page structure:
  - Historic tab shows accordion blocks, each = 1 event (event name + accuracy + profit)
  - Each accordion contains a <table> with fight rows
  - Pagination: ~3 events per page, 14 pages total

Output: backend/eda/ufc_bet_companion_picks.xlsx (two sheets: Fights, Events)

Usage:
    cd backend
    python eda/scrape_ufc_bet_companion.py
    python eda/scrape_ufc_bet_companion.py --headless false   # watch the browser
"""
from __future__ import annotations

import argparse
import re
import time
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from playwright.sync_api import sync_playwright

URL = "https://www.ufcbetcompanion.com/ai-picks"
OUTPUT = Path(__file__).parent / "ufc_bet_companion_picks.xlsx"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _parse_pct(text: str) -> float | None:
    """'62.5%' → 0.625"""
    m = re.search(r"([\d.]+)%", text or "")
    return round(float(m.group(1)) / 100, 4) if m else None


def _parse_profit(text: str) -> float | None:
    """'$12.50' / '-$7.00' → float"""
    cleaned = (text or "").strip().replace(",", "")
    m = re.search(r"(-?\$?)([\d.]+)", cleaned)
    if not m:
        return None
    val = float(m.group(2))
    neg = "-" in cleaned and "$" in cleaned
    return -val if (cleaned.startswith("-") or cleaned.startswith("−")) else val


def _parse_odds(text: str) -> int | None:
    """'+150' / '-200' → int"""
    cleaned = (text or "").replace("−", "-").strip()
    m = re.search(r"([+-]?\d+)", cleaned)
    return int(m.group(1)) if m else None


# ---------------------------------------------------------------------------
# Scraper
# ---------------------------------------------------------------------------

def scrape_page(page) -> list[dict]:
    """Scrape all event blocks + fight rows visible on the current page."""
    rows: list[dict] = []

    # Each event is an accordion block. Event headers are in div.row.border-bottom.border-top
    # Each such div is followed (as a sibling) by a div containing the table.
    # Strategy: get all event header divs and all tables in document order, then zip.

    event_headers = page.query_selector_all("div.row.border-bottom.border-top")
    tables = page.query_selector_all("table")

    if len(event_headers) != len(tables):
        print(f"  [WARN] {len(event_headers)} event headers but {len(tables)} tables — trying anyway")

    for idx, (ev_div, tbl) in enumerate(zip(event_headers, tables)):
        # Event metadata
        h2 = ev_div.query_selector("h2")
        h6s = ev_div.query_selector_all("h6")
        event_name = h2.inner_text().strip() if h2 else f"Unknown Event {idx+1}"

        accuracy_txt = h6s[0].inner_text().strip() if len(h6s) > 0 else ""
        profit_txt   = h6s[1].inner_text().strip() if len(h6s) > 1 else ""
        event_accuracy = _parse_pct(accuracy_txt)
        event_profit   = _parse_profit(profit_txt)

        # Fight rows
        trs = tbl.query_selector_all("tbody tr")
        for tr in trs:
            tds = tr.query_selector_all("td")
            if len(tds) < 5:
                continue

            fight_cell   = tds[0]
            pred_cell    = tds[1]
            correct_cell = tds[2]
            conf_cell    = tds[3]
            odds_cell    = tds[4]
            profit_cell  = tds[5] if len(tds) > 5 else None

            fight_text = fight_cell.inner_text().strip()
            fighters   = re.split(r"\s+vs\.?\s+", fight_text, flags=re.I)
            fighter_a  = fighters[0].strip() if fighters else fight_text
            fighter_b  = fighters[1].strip() if len(fighters) > 1 else None

            predicted_winner = pred_cell.inner_text().strip()

            # Correct: SVG data-icon="check" or "xmark"
            svg = correct_cell.query_selector("svg[data-icon]")
            is_correct: bool | None = None
            if svg:
                icon = (svg.get_attribute("data-icon") or "").lower()
                if icon == "check":
                    is_correct = True
                elif icon in ("xmark", "times", "x"):
                    is_correct = False

            conf_text = conf_cell.inner_text().strip()
            m = re.search(r"(\d+)", conf_text)
            confidence = int(m.group(1)) if m else None

            odds = _parse_odds(odds_cell.inner_text().strip())
            profit_val = _parse_profit(profit_cell.inner_text().strip()) if profit_cell else None

            # Fight URL (to get event ID)
            fight_link = fight_cell.query_selector("a")
            fight_url  = fight_link.get_attribute("href") if fight_link else None

            rows.append({
                "event_name":      event_name,
                "event_accuracy":  event_accuracy,
                "event_profit":    event_profit,
                "fighter_a":       fighter_a,
                "fighter_b":       fighter_b,
                "fight_text":      fight_text,
                "fight_url":       fight_url,
                "predicted_winner": predicted_winner,
                "is_correct":      is_correct,
                "confidence":      confidence,
                "odds_at_prediction": odds,
                "profit":          profit_val,
            })

    return rows


def scrape(headless: bool = True) -> list[dict]:
    all_rows: list[dict] = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless)
        page = browser.new_page(user_agent=(
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/123.0.0.0 Safari/537.36"
        ))
        page.goto(URL, wait_until="networkidle", timeout=30_000)
        time.sleep(2)

        # Click Historic tab
        historic = page.query_selector("a[data-rr-ui-event-key='Historic']")
        if not historic:
            raise RuntimeError("Could not find Historic tab on the page")
        historic.click()
        page.wait_for_selector("table", timeout=15_000)
        time.sleep(1.5)

        # Detect total pages from pagination
        page_links = page.query_selector_all("ul.pagination li.page-item a.page-link")
        page_numbers = []
        for link in page_links:
            txt = link.inner_text().strip()
            if txt.isdigit():
                page_numbers.append(int(txt))
        total_pages = max(page_numbers) if page_numbers else 1
        print(f"Total pages detected: {total_pages}")

        # Scrape page 1
        page_rows = scrape_page(page)
        all_rows.extend(page_rows)
        print(f"  Page 1: {len(page_rows)} fights ({len(all_rows)} total)")

        # Scrape pages 2..N
        for page_num in range(2, total_pages + 1):
            # Find and click the numbered pagination button
            clicked = False
            for link in page.query_selector_all("ul.pagination li.page-item a.page-link"):
                if link.inner_text().strip() == str(page_num):
                    link.click()
                    clicked = True
                    break

            if not clicked:
                print(f"  Could not find page {page_num} button, stopping early.")
                break

            # Wait for pagination active indicator to show the new page number
            try:
                page.wait_for_selector(
                    f"ul.pagination li.page-item.active span.page-link",
                    timeout=10_000,
                )
            except Exception:
                pass
            # Also wait for the table to be present with new content
            page.wait_for_selector("table tbody tr", timeout=10_000)
            time.sleep(1.2)

            page_rows = scrape_page(page)
            all_rows.extend(page_rows)
            print(f"  Page {page_num}: {len(page_rows)} fights ({len(all_rows)} total)")

        browser.close()

    return all_rows


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1E2029")
HEADER_FONT = Font(bold=True, color="FFFFFF")
GREEN_FONT  = Font(color="22C55E", bold=True)
RED_FONT    = Font(color="EF4444", bold=True)
GRAY_FONT   = Font(color="888888")


def _write_header(ws, headers: list[str]):
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def export_excel(rows: list[dict], path: Path) -> None:
    wb = openpyxl.Workbook()

    # ---- Fights sheet ----
    ws_f = wb.active
    ws_f.title = "Fights"
    fight_headers = [
        "Event", "Event Accuracy", "Event Profit ($)",
        "Fighter A", "Fighter B", "Predicted Winner",
        "Correct", "Confidence", "Odds at Prediction", "Profit ($)",
        "Fight URL",
    ]
    _write_header(ws_f, fight_headers)

    for r in rows:
        correct_str = (
            "✓" if r["is_correct"] is True
            else "✗" if r["is_correct"] is False
            else "?"
        )
        ws_f.append([
            r["event_name"],
            r["event_accuracy"],
            r["event_profit"],
            r["fighter_a"],
            r["fighter_b"],
            r["predicted_winner"],
            correct_str,
            r["confidence"],
            r["odds_at_prediction"],
            r["profit"],
            r["fight_url"],
        ])
        cell = ws_f.cell(ws_f.max_row, 7)
        if correct_str == "✓":
            cell.font = GREEN_FONT
        elif correct_str == "✗":
            cell.font = RED_FONT
        else:
            cell.font = GRAY_FONT

    col_widths = [38, 14, 16, 24, 24, 24, 9, 12, 18, 12, 55]
    for i, w in enumerate(col_widths, 1):
        ws_f.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ---- Events summary sheet ----
    ws_e = wb.create_sheet("Events")
    event_headers = [
        "Event", "Accuracy", "Event Profit ($)",
        "Fight Count", "Correct", "Wrong", "No Result",
    ]
    _write_header(ws_e, event_headers)

    from itertools import groupby
    for event_name, group in groupby(rows, key=lambda r: r["event_name"]):
        fights  = list(group)
        acc     = fights[0]["event_accuracy"]
        ep      = fights[0]["event_profit"]
        correct = sum(1 for f in fights if f["is_correct"] is True)
        wrong   = sum(1 for f in fights if f["is_correct"] is False)
        unk     = sum(1 for f in fights if f["is_correct"] is None)
        ws_e.append([event_name, acc, ep, len(fights), correct, wrong, unk])

    for i, w in enumerate([38, 12, 16, 13, 10, 10, 11], 1):
        ws_e.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wb.save(path)
    print(f"\nSaved {len(rows)} fight rows -> {path}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Scrape UFC Bet Companion historic AI picks to Excel"
    )
    parser.add_argument(
        "--headless",
        default="true",
        choices=["true", "false"],
        help="Run browser headless (default: true)",
    )
    args = parser.parse_args()

    headless = args.headless.lower() == "true"
    print(f"Starting scrape (headless={headless}) ...")
    data = scrape(headless=headless)
    print(f"\nTotal fight rows scraped: {len(data)}")
    export_excel(data, OUTPUT)
